# -*- coding: utf-8 -*-
"""
Created on Mon Aug 30 21:01:48 2021

@author: 13966
"""

import requests
import openpyxl
import matplotlib
import json
import time
def get_comments(productId,page):
    url='https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId={0}&score=0&sortType=5&page={1}&pageSize=10&isShadowSku=0&fold=1'.format(productId,page)
    headers={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36'}
    a=requests.get(url,headers=headers)
    b=a.text.replace('fetchJSON_comment98(','')
    b=b.replace(');','')
    json_data=json.loads(b)
    return json_data
def get_max_page(productId):
    d_data=get_comments(productId,0)
    return d_data['maxPage']
def get_info(productId):
    max_page=3
    lst=[]
    for page in range(1,max_page+1):
        c=get_comments(productId, page)
        c1=c['comments']
        for item in c1:
            pl=item['content']
            color=item['productColor']
            size=item['productSize']
            lst.append([pl,color,size])
        time.sleep(3)
        save(lst)
def save(lst):
    wb=openpyxl.Workbook()
    sheet=wb.active
    for item in lst:
        sheet.append(item)
    wb.save('xxx.xlsx')
    
wb=openpyxl.load_workbook('xxx.xlsx')
sheet=wb.active
rows=sheet.max_row
columns=sheet.max_column
list=[]
for i in range(1,rows+1):
    size=sheet.cell(i,3).value
    list.append(size)
dict={}
for item in list:
    dict[item]=dict.get(item,0)+1
lst1=[]
for item in dict:
    lst1.append([item,dict[item]])
print(lst1)
if __name__ == '__main__':
    productId='10024733511835'
    get_info(productId)

    