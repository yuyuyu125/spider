import openpyxl
from openpyxl.styles import Font, colors, Alignment,PatternFill,Border
from openpyxl.utils import get_column_letter
import pandas as pd
import itertools
from random import choice
import math
import numpy as np
import datetime
import os
from copy import copy
import re
import random
import warnings
warnings.filterwarnings("ignore")
#创建"人工修正"文件夹
if '人工修正' not in os.listdir():
    os.mkdir(r'./人工修正')
#获取当周周数
b=[q.split('.')[0] for q in os.listdir() if 'clr_rawWeeklyData_W'in q]
c=[q.replace('clr_rawWeeklyData_W','') for q in b if q.replace('clr_rawWeeklyData_W','').isdigit()]
week_number=max(c)
W=pd.read_excel(r'clr_rawWeeklyData_W%s.xlsx'%week_number)
W_V2=pd.read_excel(r'W%sv2.xlsx'%week_number)
W_v1=pd.read_excel(r'W%sv1.xlsx'%week_number)
column_name={
    'user2':1,
    'cityarea':2,
    'createtime':3,
    'cate':4,
    'wash_time':5,
    'name':6,
    'brand':7,
    'type':8,
    'weight1':9,
    'weight2':10,
}
double_wash={
    'U2060':'U2061',
    'U2081':'U2082',
    'U2162':'U2500',
    'U2192':'U2193',
    'U2194':'U2195',
}
double_wash_list=['U2060','U2081','U2162','U2192','U2194']
"""
排列组合顺序索引
条件：l1>=l2
例:l1=3，l2=2
输出：[(0, 1), (0, 2), (1, 2)]
"""
def shunxu_index(l1,l2):
    w = list(range(l1))
    a = list(itertools.combinations(w, l2))
    for d in a:
        for q in range(len(d) - 1):
            if d[q] > d[q + 1]:
                a.remove(d)
    return(a)

def get_list(W):
    grouped=W['user2'].groupby(W['user2'])
    grouped=list(grouped)
    t=[]
    for k in range(len(grouped)):
        if list(grouped[k][0])[0]=='U':
            t.append(grouped[k][0])
    return t

def get_double(dict_double,list):
    item=dict_double.items()
    q=[]
    for k in item:
        if k[0] in list and k[1] in list:
            q.append(k)
            print('%s和%s的数据为同一用户'%(k[0],k[1]))
    return q

def save_list(listt,namee):
    str = '\n'
    f = open("%s.txt"%namee, "w")
    f.write(str.join(listt))
    f.close()

def week_time(n):
    f = datetime.datetime(2020, 7, 13) + datetime.timedelta(weeks=n)
    g=datetime.datetime(2020,7,13)+datetime.timedelta(weeks=n+1)
    return (pd.Timestamp(f),pd.Timestamp(g))

def gouzhao_helitime(timestamp):
    return [pd.Timestamp(datetime.datetime(timestamp.year,timestamp.month,timestamp.day,12+4*kk,random.randint(0,59),random.randint(0,59))) for kk in range(3)]

def gouzhao_randomtime(timestamp,oo):
    d = []
    for iu in range(7):
        d += gouzhao_helitime(pd.Timestamp(pd.to_datetime(timestamp) + datetime.timedelta(days=iu)))
    return random.sample(d, oo)

def try_randomtime(createtime_list,washtime_list,timestamp,oo):
    try:
        createtime_list_1=createtime_list
        randomtime_list=gouzhao_randomtime(timestamp,oo)
        createtime_list+=randomtime_list
        createtime_list.sort(key=lambda x: str(x))
        for qqqq in range(len(createtime_list)-1):
            if str(createtime_list[qqqq] + pd.Timedelta(minutes=washtime_list[qqqq])) >= str(createtime_list[qqqq+1]):
                a=0/0
        return createtime_list
    except:
        try_randomtime(createtime_list_1,washtime_list,timestamp,oo)

def wash_pd (x):
    if x<=15:
        return 30
    else:
        return x

def change_v1v2(t,w,v1,v2):
    rengongpaicha=[]
    color_list=[]
    backgroud_list=[]
    more_row_color=[]
    wtime=week_time(int(week_number))
    new_w = pd.DataFrame()
    for k in t:
        print('%s开始修正中......'%k)
        if k in double_wash_list:
            k1='双洗衣机用户'+k+'('+double_wash[k]+')'
        else:
            k1=k

        user_w = w.loc[w['user2'] == k]
        user_w.reset_index(drop=True, inplace=True)

        if k in double_wash_list:
            user_v1=v1[(v1['usersktname'] == k)|(v1['usersktname'] == double_wash[k])]
            user_v1 = user_v1[user_v1['start_createtime'].notnull()]
            user_v1.sort_values('start_createtime', inplace=True, ascending=True)
        else:
            user_v1=v1.loc[v1['usersktname'] == k]
            user_v1 = user_v1[user_v1['start_createtime'].notnull()]
        user_v1.reset_index(drop=True, inplace=True)
        index_time = [qq for qq in range(len(user_v1)) if not str(wtime[0]) <= str(user_v1.loc[qq, 'start_createtime']) < str(wtime[1])]
        user_v1 = user_v1.drop(index_time, axis=0)
        user_v1.reset_index(drop=True, inplace=True)
        rpo=len(user_v1) - 1
        for er in range(rpo):
            if pd.isnull(user_v1.loc[er, 'end_createtime']):
                user_v1.loc[er + 1, 'start_createtime'] = user_v1.loc[er, 'start_createtime']
                user_v1.loc[er + 1, 'wash_time'] = math.ceil(((user_v1.loc[er + 1, 'end_createtime'] - user_v1.loc[er + 1, 'start_createtime']).total_seconds()) / 60)
                user_v1 = user_v1.drop(index=er)
        user_v1.reset_index(drop=True, inplace=True)

        if k in double_wash_list:
            user_v2 = v2[(v2['usersktname'] == k) | (v2['usersktname'] == double_wash[k])]
            user_v2 = user_v2[user_v2['dev_time'].notnull()]
            user_v2.sort_values('dev_time', inplace=True, ascending=True)
        else:
            user_v2 = v2.loc[v2['usersktname'] == k]
            user_v2 = user_v2[user_v2['dev_time'].notnull()]
        # user_v2.reset_index(drop=True, inplace=True)
        # index_time = [qq for qq in range(len(user_v2)) if not str(wtime[0]) <= str(user_v2.loc[qq, 'dev_time']) < str(wtime[1])]
        # user_v2 = user_v2.drop(index_time, axis=0)
        user_v2.reset_index(drop=True, inplace=True)

        #重量表数量相等
        if len(user_w)==len(user_v2):
            #电量表行数相等
            if len(user_w)==len(user_v1):
                for i in range(len(user_w)):
                    for d in [['weight1', 'weight1'], ['weight2', 'weight2'], ['name', 'name'], ['brand', 'brand'],['type', 'type'], ['cate', 'category']]:
                        if user_w.loc[i, d[0]] != user_v2.loc[i, d[1]] and not pd.isnull(user_v2.loc[i, d[1]]):
                            user_w.loc[i, d[0]] = user_v2.loc[i, d[1]]
                            color_list.append([(i+len(new_w)),column_name[d[0]]])
                    for d in [['createtime', 'start_createtime'], ['wash_time', 'wash_time']]:
                        if user_w.loc[i, d[0]] != user_v1.loc[i, d[1]] and not pd.isnull(user_v1.loc[i, d[1]]):
                            user_w.loc[i, d[0]] = user_v1.loc[i, d[1]]
                            color_list.append([(i + len(new_w)), column_name[d[0]]])
            #电量表行数多
            elif len(user_v1)>len(user_w):
                time_smax = []
                index_list = shunxu_index(len(user_v1), len(user_w))
                for mn in range(len(index_list)):
                    uu = index_list[mn]
                    sum_o = 0
                    for od in range(len(user_w)):
                        sum_o += abs((user_w.loc[od, 'createtime'] - user_v1.loc[uu[od], 'start_createtime']).total_seconds())
                    time_smax.append(sum_o)
                index_list_1 = index_list[time_smax.index(min(time_smax))]
                user_v1 = user_v1.loc[list(index_list_1)]
                user_v1.reset_index(drop=True, inplace=True)

                for i in range(len(user_w)):
                    for d in [['createtime', 'start_createtime'], ['wash_time', 'wash_time']]:
                        if user_w.loc[i, d[0]] != user_v1.loc[i, d[1]]and not pd.isnull(user_v1.loc[i, d[1]]):
                            user_w.loc[i, d[0]] = user_v1.loc[i, d[1]]
                            color_list.append([(i + len(new_w)), column_name[d[0]]])
                    for d in [['weight1', 'weight1'], ['weight2', 'weight2'], ['name', 'name'], ['brand', 'brand'],['type', 'type'],['cate', 'category']]:
                        if user_w.loc[i, d[0]] != user_v2.loc[i, d[1]] and not pd.isnull(user_v2.loc[i, d[1]]):
                            user_w.loc[i, d[0]] = user_v2.loc[i, d[1]]
                            color_list.append([(i + len(new_w)), column_name[d[0]]])
            #电量表无数据
            elif len(user_v1)==0:
                for i in range(len(user_w)):
                    for d in [['weight1', 'weight1'], ['weight2', 'weight2'], ['name', 'name'], ['brand', 'brand'],['type', 'type'],['cate', 'category']]:
                        if user_w.loc[i, d[0]] != user_v2.loc[i, d[1]] and not pd.isnull(user_v2.loc[i, d[1]]):
                            user_w.loc[i, d[0]] = user_v2.loc[i, d[1]]
                            color_list.append([(i + len(new_w)), column_name[d[0]]])
            #电量表行数少
            else:
                time_smax = []
                index_list = shunxu_index(len(user_w), len(user_v1))
                for mn in range(len(index_list)):
                    uu = index_list[mn]
                    sum_o = 0
                    for od in range(len(user_v1)):
                        sum_o += abs(
                            (user_w.loc[uu[od], 'createtime'] - user_v1.loc[od, 'start_createtime']).total_seconds())
                    time_smax.append(sum_o)
                index_list_1 = index_list[time_smax.index(min(time_smax))]
                for ko in range(len(index_list_1)):
                    for d in [['createtime', 'start_createtime'], ['wash_time', 'wash_time']]:
                        if user_w.loc[index_list_1[ko],d[0]] != user_v1.loc[ko, d[1]] and not pd.isnull(user_v1.loc[ko, d[1]]):
                            user_w.loc[index_list_1[ko], d[0]]=user_v1.loc[ko, d[1]]
                            color_list.append([(index_list_1[ko] + len(new_w)), column_name[d[0]]])

                for i in range(len(user_w)):
                    for d in [['weight1', 'weight1'], ['weight2', 'weight2'], ['name', 'name'], ['brand', 'brand'],['type', 'type'],['cate', 'category']]:
                        if user_w.loc[i, d[0]] != user_v2.loc[i, d[1]] and not pd.isnull(user_v2.loc[i, d[1]]):
                            user_w.loc[i, d[0]] = user_v2.loc[i, d[1]]
                            color_list.append([(i + len(new_w)), column_name[d[0]]])
        #重量表无数据
        elif len(user_v2)==0:
            # 电量表无数据
            if len(user_v1)==0:
                rengongpaicha.append('%s的重量表与电量表均无数据，需人工排查。' % k1)
            else:
                #电量表行数相等
                if len(user_w) == len(user_v1):
                    for i in range(len(user_w)):
                        for d in [['createtime', 'start_createtime'], ['wash_time', 'wash_time']]:
                            if user_w.loc[i, d[0]] != user_v1.loc[i, d[1]] and not pd.isnull(user_v1.loc[i, d[1]]):
                                user_w.loc[i, d[0]] = user_v1.loc[i, d[1]]
                                color_list.append([(i + len(new_w)), column_name[d[0]]])
                #电量表行数多
                elif len(user_v1) > len(user_w):
                    time_smax = []
                    index_list = shunxu_index(len(user_v1), len(user_w))
                    for mn in range(len(index_list)):
                        uu = index_list[mn]
                        sum_o = 0
                        for od in range(len(user_w)):
                            sum_o += abs((user_w.loc[od, 'createtime'] - user_v1.loc[uu[od], 'start_createtime']).total_seconds())
                        time_smax.append(sum_o)
                    index_list_1 = index_list[time_smax.index(min(time_smax))]
                    user_v1 = user_v1.loc[list(index_list_1)]
                    user_v1.reset_index(drop=True, inplace=True)
                    for i in range(len(user_w)):
                        for d in [['createtime', 'start_createtime'], ['wash_time', 'wash_time']]:
                            if user_w.loc[i, d[0]] != user_v1.loc[i, d[1]] and not pd.isnull(user_v1.loc[i, d[1]]):
                                user_w.loc[i, d[0]] = user_v1.loc[i, d[1]]
                                color_list.append([(i + len(new_w)), column_name[d[0]]])
                # 电量表行数少
                else:
                    time_smax = []
                    index_list = shunxu_index(len(user_w), len(user_v1))
                    for mn in range(len(index_list)):
                        uu = index_list[mn]
                        sum_o = 0
                        for od in range(len(user_v1)):
                            sum_o += abs((user_w.loc[uu[od], 'createtime'] - user_v1.loc[od, 'start_createtime']).total_seconds())
                        time_smax.append(sum_o)
                    index_list_1 = index_list[time_smax.index(min(time_smax))]
                    for ko in range(len(index_list_1)):
                        for d in [['createtime', 'start_createtime'], ['wash_time', 'wash_time']]:
                            if user_w.loc[index_list_1[ko], d[0]] != user_v1.loc[ko, d[1]] and not pd.isnull(user_v1.loc[ko, d[1]]):
                                user_w.loc[index_list_1[ko], d[0]] = user_v1.loc[ko, d[1]]
                                color_list.append([(index_list_1[ko] + len(new_w)), column_name[d[0]]])
                rengongpaicha.append('%s的重量表无数据，电量表有数据，已经过程序检查，但需人工排查。'%k)
        #重量表行数少（问卷重复）
        elif len(user_v2)<len(user_w):
            time_smax = []
            index_list = shunxu_index(len(user_w), len(user_v2))
            for mn in range(len(index_list)):
                uu = index_list[mn]
                sum_o = 0
                for od in range(len(user_v2)):
                    sum_o += abs((user_w.loc[uu[od], 'createtime'] - user_v2.loc[od, 'dev_time']).total_seconds())
                time_smax.append(sum_o)
            index_list_1 = index_list[time_smax.index(min(time_smax))]#需要匹配行的索引
            index_list_2=list(set(list(range(len(user_w))))-set(index_list_1))
            for ll in index_list_2:
                backgroud_list.append(len(new_w)+ll)
            for qq in range(len(user_v2)):
                for d in [['weight1', 'weight1'], ['weight2', 'weight2'], ['name', 'name'], ['brand', 'brand'],['type', 'type'],['cate', 'category']]:
                    if user_w.loc[index_list_1[qq], d[0]] != user_v2.loc[qq, d[1]] and not pd.isnull(user_v2.loc[qq, d[1]]):
                        user_w.loc[index_list_1[qq], d[0]] = user_v2.loc[qq, d[1]]
                        color_list.append([(index_list_1[qq] + len(new_w)), column_name[d[0]]])
            #重量表和电量表行数相等
            if len(user_v1)==len(user_v2):
                for qq in range(len(user_v1)):
                    for d in [['createtime', 'start_createtime'], ['wash_time', 'wash_time']]:
                        if user_w.loc[index_list_1[qq],d[0]] != user_v1.loc[qq, d[1]] and not pd.isnull(user_v1.loc[qq, d[1]]):
                            user_w.loc[index_list_1[qq], d[0]]=user_v1.loc[qq, d[1]]
                            color_list.append([(index_list_1[qq] + len(new_w)), column_name[d[0]]])
            # 重量表比电量表行数少
            elif len(user_v1)>len(user_v2):
                time_smax = []
                index_list = shunxu_index(len(user_v1), len(user_v2))
                for mn in range(len(index_list)):
                    uu = index_list[mn]
                    sum_o = 0
                    for od in range(len(user_v2)):
                        sum_o += abs((user_w.loc[index_list_1[od], 'createtime'] - user_v1.loc[uu[od], 'start_createtime']).total_seconds())
                    time_smax.append(sum_o)
                index_list_1_1 = index_list[time_smax.index(min(time_smax))]
                for qq in range(len(user_v2)):
                    for d in [['createtime', 'start_createtime'], ['wash_time', 'wash_time']]:
                        if user_w.loc[index_list_1[qq], d[0]] != user_v1.loc[index_list_1_1[qq], d[1]] and not pd.isnull(user_v1.loc[index_list_1_1[qq], d[1]]):
                            user_w.loc[index_list_1[qq], d[0]] = user_v1.loc[index_list_1_1[qq], d[1]]
                            color_list.append([(index_list_1[qq] + len(new_w)), column_name[d[0]]])
            # 重量表比电量表行数多
            elif len(user_v1)!=0:
                time_smax = []
                index_list = shunxu_index(len(user_v2), len(user_v1))
                for mn in range(len(index_list)):
                    uu = index_list[mn]
                    sum_o = 0
                    for od in range(len(user_v1)):
                        sum_o += abs((user_w.loc[index_list_1[uu[od]], 'createtime'] - user_v1.loc[od, 'start_createtime']).total_seconds())
                    time_smax.append(sum_o)
                index_list_1_1 = index_list[time_smax.index(min(time_smax))]
                for qq in range(len(user_v1)):
                    for d in [['createtime', 'start_createtime'], ['wash_time', 'wash_time']]:
                        if user_w.loc[index_list_1[index_list_1_1[qq]], d[0]] != user_v1.loc[qq, d[1]] and not pd.isnull(user_v1.loc[qq, d[1]]):
                            user_w.loc[index_list_1[index_list_1_1[qq]], d[0]] = user_v1.loc[qq, d[1]]
                            color_list.append([(index_list_1[index_list_1_1[qq]] + len(new_w)), column_name[d[0]]])
            rengongpaicha.append('%s的重量表数量少，已经过程序检查，但需人工排查。' % k1)
        # 重量表行数多（缺少问卷）
        else:
            time_smax = []
            index_list = shunxu_index(len(user_v2), len(user_w))
            for mn in range(len(index_list)):
                uu = index_list[mn]
                sum_o = 0
                for od in range(len(user_w)):
                    sum_o += abs((user_v2.loc[uu[od], 'dev_time'] - user_w.loc[od, 'createtime']).total_seconds())
                time_smax.append(sum_o)
            index_list_min = index_list[time_smax.index(min(time_smax))]
            user_w.index=list(index_list_min)
            index_list_min_yu=[lop for lop in list(range(len(user_v2))) if lop not in index_list_min]
            for yu_index in index_list_min_yu:
                user_w.loc[yu_index]=user_w.loc[index_list_min[0]]
                more_row_color.append(len(new_w)+yu_index)
            user_w.sort_index(inplace=True)
            for i in range(len(user_w)):
                for d in [['weight1', 'weight1'], ['weight2', 'weight2'], ['name', 'name'], ['brand', 'brand'],['type', 'type'], ['cate', 'category']]:
                    if user_w.loc[i, d[0]] != user_v2.loc[i, d[1]] and not pd.isnull(user_v2.loc[i, d[1]]):
                        user_w.loc[i, d[0]] = user_v2.loc[i, d[1]]
                        color_list.append([(i + len(new_w)), column_name[d[0]]])
            #电量表行数相等
            if len(user_v1)==len(user_w):
                for i in range(len(user_w)):
                    for d in [['createtime', 'start_createtime'], ['wash_time', 'wash_time']]:
                        if user_w.loc[i, d[0]] != user_v1.loc[i, d[1]] and not pd.isnull(user_v1.loc[i, d[1]]):
                            user_w.loc[i, d[0]] = user_v1.loc[i, d[1]]
                            color_list.append([(i + len(new_w)), column_name[d[0]]])
            # 电量表行数多
            elif len(user_v1)>len(user_w):
                time_smax = []
                index_list = shunxu_index(len(user_v1), len(user_w))
                for mn in range(len(index_list)):
                    uu = index_list[mn]
                    sum_o = 0
                    for od in range(len(user_w)):
                        sum_o += abs((user_w.loc[od, 'createtime'] - user_v1.loc[uu[od], 'start_createtime']).total_seconds())
                    time_smax.append(sum_o)
                index_list_1 = index_list[time_smax.index(min(time_smax))]
                user_v1 = user_v1.loc[list(index_list_1)]
                user_v1.reset_index(drop=True, inplace=True)
                for i in range(len(user_w)):
                    for d in [['createtime', 'start_createtime'], ['wash_time', 'wash_time']]:
                        if user_w.loc[i, d[0]] != user_v1.loc[i, d[1]] and not pd.isnull(user_v1.loc[i, d[1]]):
                            user_w.loc[i, d[0]] = user_v1.loc[i, d[1]]
                            color_list.append([(i + len(new_w)), column_name[d[0]]])
            # 电量表行数少
            elif len(user_v1)!=0:
                time_smax = []
                index_list = shunxu_index(len(user_w), len(user_v1))
                for mn in range(len(index_list)):
                    uu = index_list[mn]
                    sum_o = 0
                    for od in range(len(user_v1)):
                        sum_o += abs((user_w.loc[uu[od], 'createtime'] - user_v1.loc[od, 'start_createtime']).total_seconds())
                    time_smax.append(sum_o)
                index_list_1 = index_list[time_smax.index(min(time_smax))]
                for ko in range(len(index_list_1)):
                    for d in [['createtime', 'start_createtime'], ['wash_time', 'wash_time']]:
                        if user_w.loc[index_list_1[ko], d[0]] != user_v1.loc[ko, d[1]] and not pd.isnull(user_v1.loc[ko, d[1]]):
                            user_w.loc[index_list_1[ko], d[0]] = user_v1.loc[ko, d[1]]
                            color_list.append([(index_list_1[ko] + len(new_w)), column_name[d[0]]])
            rengongpaicha.append('%s的重量表数量多，已经过程序检查，但需人工排查。'%k1)
        #去除多余没意义文字
        for lx in range(len(user_w)):
            if '分钟' in str(user_w.loc[lx, 'wash_time']):
                user_w.loc[lx, 'wash_time'] = float(str(user_w.loc[lx, 'wash_time']).replace('分钟', ''))
                color_list.append([(lx + len(new_w)), column_name['wash_time']])
            elif '小时' in str(user_w.loc[lx, 'wash_time']):
                user_w.loc[lx, 'wash_time'] = float(str(user_w.loc[lx, 'wash_time']).replace('小时', ''))*60
                color_list.append([(lx + len(new_w)), column_name['wash_time']])

        user_w['wash_time']=user_w['wash_time'].apply(lambda x:wash_pd(x))
        #重量修正
        if len(user_w) > 1:
            for o in range(len(user_w) - 1):
                if (o + len(new_w)) not in backgroud_list:
                    for o1 in range(o+1,len(user_w)):
                        if not pd.isnull(user_w.loc[o, 'weight2']) and not pd.isnull(user_w.loc[o1, 'weight1']) and (o1 + len(new_w)) not in backgroud_list:
                            if (-5<=user_w.loc[o, 'weight2']-user_w.loc[o1, 'weight1']<0 or 1<(user_w.loc[o, 'weight2']-user_w.loc[o1,'weight1'])<=5) and user_w.loc[o, 'name']==user_w.loc[o1, 'name']:
                                user_w.loc[o1, 'weight1'] = user_w.loc[o, 'weight2']
                                color_list.append([(o1 + len(new_w)), column_name['weight1']])
                                break
            #重量重复标黄
            for mi in range(len(user_w) - 1):
                if (mi + len(new_w)) not in backgroud_list:
                    for mi1 in range(mi+1,len(user_w)):
                        #参数
                        if 0<=(user_w.loc[mi, 'weight1']+user_w.loc[mi, 'weight2']-user_w.loc[(mi1), 'weight1']-user_w.loc[(mi1), 'weight2'])<=2 and  user_w.loc[mi, 'name']==user_w.loc[mi1, 'name'] and (mi1 + len(new_w)) not in backgroud_list:
                            backgroud_list.append(mi1 + len(new_w))


        #时间修正
            rtb = []
            rty = []
            lingyi=[]
            max_notback_index=max([bb for bb in range(len(user_w)) if (len(new_w)+bb) not in backgroud_list])
            for uop in range(len(user_w)):
                rty.append(uop + 1)
                rty.append(-uop - 1)
            for m in range(len(user_w)-1):

                if (m + len(new_w)) not in backgroud_list:
                    gp=1
                    while (m +gp+ len(new_w)) in backgroud_list:
                        gp+=1
                    if m+gp+1<=len(user_w):
                        #参数
                        if  (abs((user_w.loc[m, 'createtime']-user_w.loc[(m+gp), 'createtime']).total_seconds())<=60 or user_w.loc[m, 'quote_time']==user_w.loc[(m+gp), 'quote_time']) and user_w.loc[m, 'name']!=user_w.loc[(m+gp), 'name']:
                            if user_w.loc[m, 'createtime']!=user_w.loc[(m + gp), 'createtime']:
                                user_w.loc[(m + gp), 'createtime'] = user_w.loc[m, 'createtime']
                                color_list.append([m+gp+len(new_w),column_name['createtime']])

                            if (len(lingyi)==0 or lingyi[-1]==1) and user_w.loc[m, 'wash_time'] in rtb:
                                for kkkk in rty:
                                    if (kkkk + user_w.loc[m, 'wash_time']) not in rtb:
                                        user_w.loc[m, 'wash_time'] += kkkk
                                        color_list.append([m + len(new_w), column_name['wash_time']])
                                        break
                                rtb.append(user_w.loc[m, 'wash_time'])

                            if  user_w.loc[(m + gp), 'wash_time']!= user_w.loc[m, 'wash_time']:
                                user_w.loc[(m + gp), 'wash_time']=user_w.loc[m, 'wash_time']
                                color_list.append([m + gp + len(new_w), column_name['wash_time']])
                            lingyi.append(0)
                        else:
                            #debug
                            if (len(lingyi)==0 or lingyi[-1]==1) and user_w.loc[m, 'wash_time'] in rtb:
                                for kkkk in rty:
                                    if (kkkk+user_w.loc[m, 'wash_time']) not in rtb:
                                        user_w.loc[m, 'wash_time']+=kkkk
                                        color_list.append([m + len(new_w), column_name['wash_time']])
                                        break
                            rtb.append(user_w.loc[m, 'wash_time'])
                            #最后一个washtime
                            if m+gp==max_notback_index and user_w.loc[m + gp, 'wash_time'] in rtb:
                                for kkk in rty:
                                    if (kkk+user_w.loc[m + gp, 'wash_time']) not in rtb:
                                        user_w.loc[m + gp, 'wash_time']+=kkk
                                        color_list.append([m +gp+ len(new_w), column_name['wash_time']])
                                        break
                            lingyi.append(1)
                            if str(user_w.loc[m, 'createtime'] + pd.Timedelta(minutes=user_w.loc[m, 'wash_time'])) > str(user_w.loc[m+gp, 'createtime']):
                                # 参数
                                user_w.loc[m + gp, 'createtime']=user_w.loc[m, 'createtime'] + pd.Timedelta(minutes=(user_w.loc[m, 'wash_time']+10))
                                color_list.append([m + gp + len(new_w), column_name['createtime']])



        #日期超出范围进行修正
        for dfn in range(len(user_w)):
            if not str(wtime[0])<=str(user_w.loc[dfn,'createtime'])<str(wtime[1]) and (dfn + len(new_w)) not in backgroud_list:
                fei_back_index=[uu for uu in range(len(user_w)) if (len(new_w)+uu) not in backgroud_list]
                fei_back_wash_time_list=[list(user_w['wash_time'])[oo] for oo in fei_back_index]
                fei_back_createtime_list=[list(user_w['createtime'])[oo] for oo in fei_back_index]
                for aa in range(1,len(fei_back_index)):
                    if fei_back_createtime_list[-aa]!=fei_back_createtime_list[-aa-1]+pd.Timedelta(minutes=(fei_back_wash_time_list[-aa-1]+10)):
                        break
                for pp in range(aa):
                    del fei_back_createtime_list[-1]
                #同次多care
                if len(fei_back_createtime_list)!=len(list(set(fei_back_createtime_list))):
                    rengongpaicha.append('%s经过程序修正后的日期错误，需人工修正日期。' % k1)
                else:
                    new_createtime_list=try_randomtime(fei_back_createtime_list,fei_back_wash_time_list,wtime[0],(len(fei_back_wash_time_list)-len(fei_back_createtime_list)))
                    for ppp in range(len(fei_back_index)):
                        user_w.loc[fei_back_index[ppp],'createtime']=new_createtime_list[ppp]
                        color_list.append([fei_back_index[ppp]+ len(new_w), column_name['createtime']])
                    break
        #对整段标红行进行'quote_time'字段修改
        for hh in range(len(user_w)):
            if (len(new_w)+hh) in more_row_color:
                user_w.loc[hh, 'quote_time'] = user_w.loc[hh, 'createtime']
        print('%s已修正成功......'%k)
        new_w = pd.concat([new_w, user_w], axis=0)
    new_w.reset_index(drop=True, inplace=True)
    print('全部数据已修正完成......')
    save_list(rengongpaicha,'./人工修正/人工修正')
    return (new_w,color_list,backgroud_list,more_row_color)

def run():
    W_list=get_list(W)
    op=get_double(double_wash,W_list)
    NEW=change_v1v2(W_list,W,W_v1,W_V2)
    name='W%s程序修改结果.xlsx'%week_number
    NEW[0].to_excel(name,index=False)
    print('正在修正excel格式......')
    a=openpyxl.load_workbook(name)
    sheet=a['Sheet1']
    alignment_right=Alignment(horizontal='left',vertical='center')
    ffo=Font(name='等线', size=11)
    border = Border()
    ff=Font(name='等线', size=11, color="FF0000")
    yello_fill = PatternFill("solid", fgColor="FFFF00")
    for ky in range(sheet.max_column):
        sheet.cell(row=1,column=(ky+1)).border = border
        sheet.cell(row=1,column=(ky+1)).alignment=alignment_right
    for zx in range(sheet.max_column):
        for un in range(sheet.max_row):
            sheet.cell(row=(un+1),column=(zx+1)).font=ffo
    for ij in NEW[1]:
        #黄行无红字
        if ij[0] not in NEW[2]:
            sheet.cell(row=(ij[0]+2), column=(ij[1])).font = ff
    for ui in NEW[2]:
        for k in range(sheet.max_column):
            sheet.cell(row=(ui + 2), column=(k+1)).fill = yello_fill
    for kip in NEW[3]:
        for k in range(sheet.max_column):
            sheet.cell(row=(kip + 2), column=(k + 1)).font = ff
    a.save(name)
    print('excel格式已修正完毕......')
    #合并excel
    print('正在合并excel......')
    xin = openpyxl.Workbook()
    xinws = xin.active
    b= openpyxl.load_workbook(r'clr_rawWeeklyData_W%s.xlsx'%week_number)
    sheet_b=b['Sheet1']
    row_b=sheet_b.max_row
    col_b=sheet_b.max_column
    for r in range(row_b):
        if 'U' in str(sheet_b.cell(row=(r + 1), column=1).value):
            break
        for c in range(col_b):
            if str(sheet_b.cell(row=(r + 1), column=(c + 1)).value)=='NULL':
                xinws.cell(row=(r + 1), column=(c + 1)).value=''
            else:
                xinws.cell(row=(r + 1), column=(c + 1)).value = sheet_b.cell(row=(r + 1), column=(c + 1)).value
            xinws.cell(row=(r + 1), column=(c + 1)).font = ffo
            xinws.cell(row=(r + 1), column=(c + 1)).fill = copy(sheet_b.cell(row=(r + 1), column=(c + 1)).fill)
            xinws.cell(row=(r + 1), column=(c + 1)).border = copy(sheet_b.cell(row=(r + 1), column=(c + 1)).border)
            xinws.cell(row=(r + 1), column=(c + 1)).alignment = copy(sheet_b.cell(row=(r + 1), column=(c + 1)).alignment)
            xinws.cell(row=(r + 1), column=(c + 1)).protection = copy(sheet_b.cell(row=(r + 1), column=(c + 1)).protection)
            xinws.cell(row=(r + 1), column=(c + 1)).number_format = copy(sheet_b.cell(row=(r + 1), column=(c + 1)).number_format)
    a = openpyxl.load_workbook(name)
    sheet = a['Sheet1']
    row=sheet.max_row
    col=sheet.max_column
    for ro in range(row-1):
        for co in range(col):
            xinws.cell(row=(r+ro+1), column=(co + 1)).value = sheet.cell(row=(ro + 2), column=(co + 1)).value
            xinws.cell(row=(r+ro+1), column=(co + 1)).font = copy(sheet.cell(row=(ro + 2), column=(co + 1)).font)
            xinws.cell(row=(r+ro+1), column=(co + 1)).fill = copy(sheet.cell(row=(ro + 2), column=(co + 1)).fill)
            xinws.cell(row=(r+ro+1), column=(co + 1)).border = copy(sheet.cell(row=(ro + 2), column=(co + 1)).border)
            xinws.cell(row=(r+ro+1), column=(co+ 1)).alignment = copy(sheet.cell(row=(ro + 2), column=(co + 1)).alignment)
            xinws.cell(row=(r+ro+1), column=(co + 1)).protection = copy(sheet.cell(row=(ro + 2), column=(co + 1)).protection)
            xinws.cell(row=(r+ro+1), column=(co + 1)).number_format = copy(sheet.cell(row=(ro + 2), column=(co + 1)).number_format)
    dims = {}
    # 遍历表格数据，获取自适应列宽数据
    for row in xinws.rows:
        for cell in row:
            if cell.value:
                cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
    xinws.column_dimensions[get_column_letter(column_name['createtime'])].width = dims[column_name['createtime']]
    xin.save(name)
    xin.save(r'./人工修正/clr_rawWeeklyData_W%s.xlsx'%week_number)
    print('excel已合并完毕......')
run()